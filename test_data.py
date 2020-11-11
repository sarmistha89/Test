import pytest
import logging
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from BaseClass import BaseClass

@pytest.mark.usefixtures("setup")
class Testdatabase(BaseClass):
    '''
    def test_database(self):
        try:
            book=openpyxl.load_workbook("/home/pi/Documents/computer Database test/computerlist.xlsx")
            sheet=book.active
            row=sheet.max_row
            print(row)
            col=sheet.max_column
            for i in range (2,row+1):
                data=sheet.cell(row=i,column=1).value
                self.driver.find_element_by_id("searchbox").send_keys(data)
                self.driver.find_element_by_id("searchsubmit").click()
                itemlist=[]
                element= self.driver.find_elements_by_xpath("//td/a[contains(@href,'/computers/')]")
                if element == []:
                    print("The data is not available")
                                                            
                else:
                    for elem in element:
                        itemlist.append(elem.text)
                        #print(itemlist)
                        print("the item", elem.text, "is available")
                self.driver.find_element_by_id("searchbox").clear()
                
              #WebDriverWait(self.driver, 10).until(EC.visibility_of((data)))
        except Exception as e:
            print(e)'''
             
    def test_addcomputer(self):
        n=input("Enter the computer name")
        self.driver.find_element_by_id("searchbox").send_keys(n)
        self.driver.find_element_by_id("searchsubmit").click()
        self.driver.find_element_by_id("add").click()
        self.driver.find_element_by_name("name").send_keys(n)
        select=Select(self.driver.find_element_by_id('company'))
        select.select_by_value('2')
        self.driver.find_element_by_xpath("//input[@value='Create this computer']").click()
        confirmation=self.driver.find_element_by_xpath("//div[@class='alert-message warning']").text
        assert "Done !" in confirmation
        self.driver.find_element_by_id("searchbox").send_keys(n)
        self.driver.find_element_by_id("searchsubmit").click()
        log=self.logging()
        log.error("Added data is not displaying")
                
        
            
        '''
        book=openpyxl.load_workbook("/home/pi/Documents/computer Database test/computerlist.xlsx")
        sheet=book.active
        row=sheet.max_row
        print(row)
        col=sheet.max_column
        for i in range (1,row+1):
            data=sheet.cell(row=i,column=1).value
            self.driver.find_element_by_id("searchbox").send_keys(data)
            self.driver.find_element_by_id("searchsubmit").click()
            try:
                WebDriverWait(self.driver, 10).until(EC.visibility_of((data)))
                print("the item", data, "is available")
                self.driver.find_element_by_id("searchbox").clear()
                
              
            except:
                print("The data is not available")'''